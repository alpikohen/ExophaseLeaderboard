import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook, load_workbook

url_general = "https://www.exophase.com/leaderboard/"
response = requests.get(url_general)
general_data = response.content
soup = BeautifulSoup(general_data, "html.parser")
general_user_number = soup.find_all("span", {"class":"big_number"})
message = str(general_user_number[1]).split('>')
message2 = message[1].split('<')
general_total = int(message2[0].replace(',', ''))

url_psn = "https://www.exophase.com/psn/leaderboard/"
response = requests.get(url_psn)
psn_data = response.content
soup = BeautifulSoup(psn_data, "html.parser")
psn_user_number = soup.find_all("span", {"class":"big_number"})
message = str(psn_user_number[1]).split('>')
message2 = message[1].split('<')
psn_total = int(message2[0].replace(',', ''))

url_xbox = "https://www.exophase.com/xbox/leaderboard/"
response = requests.get(url_xbox)
xbox_data = response.content
soup = BeautifulSoup(xbox_data, "html.parser")
xbox_user_number = soup.find_all("span", {"class":"big_number"})
message = str(xbox_user_number[1]).split('>')
message2 = message[1].split('<')
xbox_total = int(message2[0].replace(',', ''))

url_steam = "https://www.exophase.com/steam/leaderboard/"
response = requests.get(url_steam)
steam_data = response.content
soup = BeautifulSoup(steam_data, "html.parser")
steam_user_number = soup.find_all("span", {"class":"big_number"})
message = str(steam_user_number[1]).split('>')
message2 = message[1].split('<')
steam_total = int(message2[0].replace(',', ''))

url_ea = "https://www.exophase.com/origin/leaderboard/"
response = requests.get(url_ea)
ea_data = response.content
soup = BeautifulSoup(ea_data, "html.parser")
ea_user_number = soup.find_all("span", {"class":"big_number"})
message = str(ea_user_number[1]).split('>')
message2 = message[1].split('<')
ea_total = int(message2[0].replace(',', ''))

url_retro = "https://www.exophase.com/retro/leaderboard/"
response = requests.get(url_retro)
retro_data = response.content
soup = BeautifulSoup(retro_data, "html.parser")
retro_user_number = soup.find_all("span", {"class":"big_number"})
message = str(retro_user_number[1]).split('>')
message2 = message[1].split('<')
retro_total = int(message2[0].replace(',', ''))

url_gplay = "https://www.exophase.com/android/leaderboard/"
response = requests.get(url_gplay)
gplay_data = response.content
soup = BeautifulSoup(gplay_data, "html.parser")
gplay_user_number = soup.find_all("span", {"class":"big_number"})
message = str(gplay_user_number[1]).split('>')
message2 = message[1].split('<')
gplay_total = int(message2[0].replace(',', ''))

url_gog = "https://www.exophase.com/gog/leaderboard/"
response = requests.get(url_gog)
gog_data = response.content
soup = BeautifulSoup(gog_data, "html.parser")
gog_user_number = soup.find_all("span", {"class":"big_number"})
message = str(gog_user_number[1]).split('>')
message2 = message[1].split('<')
gog_total = int(message2[0].replace(',', ''))

url_ubisoft = "https://www.exophase.com/uplay/leaderboard/"
response = requests.get(url_ubisoft)
ubisoft_data = response.content
soup = BeautifulSoup(ubisoft_data, "html.parser")
ubisoft_user_number = soup.find_all("span", {"class":"big_number"})
message = str(ubisoft_user_number[1]).split('>')
message2 = message[1].split('<')
ubisoft_total = int(message2[0].replace(',', ''))

url_epic = "https://www.exophase.com/epic/leaderboard/"
response = requests.get(url_epic)
epic_data = response.content
soup = BeautifulSoup(epic_data, "html.parser")
epic_user_number = soup.find_all("span", {"class":"big_number"})
message = str(epic_user_number[1]).split('>')
message2 = message[1].split('<')
epic_total = int(message2[0].replace(',', ''))


url_user_general = "https://www.exophase.com/user/alpikohen/"
response = requests.get(url_user_general)
general_user_data = response.content
soup = BeautifulSoup(general_user_data, "html.parser")
general_ranking = soup.find_all("span", {"class":"global-ranking tippy"})
message = str(general_ranking).split('>')
message2 = message[3].split('<')
general_mine = int(message2[0].replace(',', ''))

url_user_psn = "https://www.exophase.com/psn/user/alpikohen/"
response = requests.get(url_user_psn)
psn_user_data = response.content
soup = BeautifulSoup(psn_user_data, "html.parser")
psn_ranking = soup.find_all("span", {"class":"global-ranking tippy mb-1"})
message = str(psn_ranking).split('>')
message2 = message[3].split('<')
psn_mine = int(message2[0].replace(',', ''))

url_user_xbox = "https://www.exophase.com/xbox/user/alpikohen/"
response = requests.get(url_user_xbox)
xbox_user_data = response.content
soup = BeautifulSoup(xbox_user_data, "html.parser")
xbox_ranking = soup.find_all("span", {"class":"global-ranking tippy"})
message = str(xbox_ranking).split('>')
message2 = message[3].split('<')
xbox_mine = int(message2[0].replace(',', ''))

url_user_steam= "https://www.exophase.com/steam/user/alpikohen/"
response = requests.get(url_user_steam)
steam_user_data = response.content
soup = BeautifulSoup(steam_user_data, "html.parser")
steam_ranking = soup.find_all("span", {"class":"global-ranking tippy"})
message = str(steam_ranking).split('>')
message2 = message[3].split('<')
steam_mine = int(message2[0].replace(',', ''))

url_user_ea = "https://www.exophase.com/origin/user/alpikohenn/"
response = requests.get(url_user_ea)
ea_user_data = response.content
soup = BeautifulSoup(ea_user_data, "html.parser")
ea_ranking = soup.find_all("span", {"class":"global-ranking tippy"})
message = str(ea_ranking).split('>')
message2 = message[3].split('<')
ea_mine = int(message2[0].replace(',', ''))

url_user_retro = "https://www.exophase.com/retro/user/alpikohen/"
response = requests.get(url_user_retro)
retro_user_data = response.content
soup = BeautifulSoup(retro_user_data, "html.parser")
retro_ranking = soup.find_all("span", {"class":"global-ranking tippy"})
message = str(retro_ranking).split('>')
message2 = message[3].split('<')
retro_mine = int(message2[0].replace(',', ''))

url_user_gplay = "https://www.exophase.com/android/user/alpikohen/"
response = requests.get(url_user_gplay)
gplay_user_data = response.content
soup = BeautifulSoup(gplay_user_data, "html.parser")
gplay_ranking = soup.find_all("span", {"class":"global-ranking tippy"})
message = str(gplay_ranking).split('>')
message2 = message[3].split('<')
gplay_mine = int(message2[0].replace(',', ''))

url_user_gog = "https://www.exophase.com/gog/user/alpikohen/"
response = requests.get(url_user_gog)
gog_user_data = response.content
soup = BeautifulSoup(gog_user_data, "html.parser")
gog_ranking = soup.find_all("span", {"class":"global-ranking tippy"})
message = str(gog_ranking).split('>')
message2 = message[3].split('<')
gog_mine = int(message2[0].replace(',', ''))

url_user_ubisoft = "https://www.exophase.com/uplay/user/alpikohen/"
response = requests.get(url_user_ubisoft)
ubisoft_user_data = response.content
soup = BeautifulSoup(ubisoft_user_data, "html.parser")
ubisoft_ranking = soup.find_all("span", {"class":"global-ranking tippy"})
message = str(ubisoft_ranking).split('>')
message2 = message[3].split('<')
ubisoft_mine = int(message2[0].replace(',', ''))

url_user_epic = "https://www.exophase.com/epic/user/alpikohen/"
response = requests.get(url_user_epic)
epic_user_data = response.content
soup = BeautifulSoup(epic_user_data, "html.parser")
epic_ranking = soup.find_all("span", {"class":"global-ranking tippy"})
message = str(epic_ranking).split('>')
message2 = message[3].split('<')
epic_mine = int(message2[0].replace(',', ''))

wb = load_workbook("Achievement_Track.xlsx")
ws = wb["Sayfa1"]
ws['C3']  = psn_mine
ws['C4']  = xbox_mine
ws['C5']  = steam_mine
ws['C6']  = ea_mine
ws['C7']  = retro_mine
ws['C8']  = gplay_mine
ws['C9']  = gog_mine
ws['C10'] = ubisoft_mine
ws['C11'] = epic_mine
ws['C12'] = general_mine

ws['D3']  = psn_total
ws['D4']  = xbox_total
ws['D5']  = steam_total
ws['D6']  = ea_total
ws['D7']  = retro_total
ws['D8']  = gplay_total
ws['D9']  = gog_total
ws['D10'] = ubisoft_total
ws['D11'] = epic_total
ws['D12'] = general_total
wb.save("Achievement_Track.xlsx")